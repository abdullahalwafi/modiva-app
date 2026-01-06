from django.shortcuts import render, redirect
from django.views.generic import ListView,DetailView, TemplateView
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.http import request,HttpResponseRedirect
from django.db.models import Q

from modules.vitamin.models import Distribusiobat

from modules.core.core_libs import *
from django.contrib.auth.models import Group

from django.contrib.auth.mixins import PermissionRequiredMixin
import logging

logger = logging.getLogger(__name__)
from django.forms import formset_factory
from modules.uman.decorators  import group_must_have_permission
from django.utils.decorators import method_decorator
from django.views import View
from datetime import date
from django.core.exceptions import ObjectDoesNotExist

from .forms import *

from django.utils import timezone

import os
from django.conf import settings
from django.http import FileResponse, Http404
from openpyxl import load_workbook
from io import BytesIO

import pandas as pd
from django.shortcuts import redirect
from django.contrib import messages
from .models import Vitamin, Satuan, Stokobat  # Adjust model names to match your actual models

from django.db.models import F

from django.db.models.deletion import ProtectedError
from collections import defaultdict

# ----------------referensi DisObat--------------------

@method_decorator([group_must_have_permission], name='dispatch')
class DisObatListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Distribusiobat
    permission_required = 'vitamin.view_distribusiobat'

    def get_template_names(self):
        if self.request.htmx:
            return ["vitamin/puskesmas/disobat/disobat_list.html"]
        else:
            return ["vitamin/puskesmas/disobat/disobat.html"]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["numlist"] = (int(self.request.GET.get("page", 1)) - 1)
        context["q"] = self.request.GET.get("q", '')
        context["title"] = "Daftar Distribusi Obat"

        puskesmas_id = self.request.GET.get('id', '0')
        app_group = Group.objects.get(name='administrator')

        if not self.request.user.is_superuser and app_group not in self.request.user.groups.all():
            try:
                puskesmas_id = Puskesmas.objects.get(kode=self.request.user).id
            except Exception:
                puskesmas_id = 0
            qs = self.model.objects.filter(puskesmas_id=puskesmas_id).count()
        else:
            qs = self.model.objects.all().count()

        context["count2"] = qs
        return context

    def get_queryset(self):
        app_group = Group.objects.get(name='administrator')
        app_group2 = Group.objects.get(name='Puskesmas')
        app_group3 = Group.objects.get(name='Sekolah')
        search_query = self.request.GET.get("q", "").strip()

        qs = self.model.objects.select_related("stokobat", "puskesmas", "sekolah")

        if not self.request.user.is_superuser and app_group not in self.request.user.groups.all():
            if app_group2 in self.request.user.groups.all():
                try:
                    puskesmas_id = Puskesmas.objects.get(kode=self.request.user).id
                except Exception:
                    puskesmas_id = 0
                qs = qs.filter(puskesmas_id=puskesmas_id).order_by('-tgl_terima', '-id')
                if search_query:
                    qs = qs.filter(
                        Q(tgl_kirim__icontains=search_query) |
                        Q(tgl_terima__icontains=search_query) |
                        Q(sekolah__nama__icontains=search_query)
                    )
            elif app_group3 in self.request.user.groups.all():
                try:
                    sekolah_id = Sekolah.objects.get(kode=self.request.user).id
                except Exception:
                    sekolah_id = 0
                qs = qs.filter(sekolah_id=sekolah_id).order_by('-tgl_terima', '-id')
            else:
                qs = qs.none()
        else:
            qs = qs.order_by('-tgl_terima', '-id')
            if search_query:
                qs = qs.filter(
                    Q(tgl_kirim__icontains=search_query) |
                    Q(tgl_terima__icontains=search_query) |
                    Q(puskesmas__nama__icontains=search_query) |
                    Q(sekolah__nama__icontains=search_query)
                )
        return qs


class DisObatCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    model = Distribusiobat
    template_name = 'vitamin/puskesmas/disobat/disobat_form.html'
    form_class = DisObatForm
    success_url = reverse_lazy('vitamin:disobat-list')
    permission_required = 'vitamin.add_distribusiobat'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Tambah Distribusi Obat"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # ✅ Pass the logged-in user to the form
        return kwargs

    def form_valid(self, form):
        with transaction.atomic():
            obj = form.save(commit=False)
            stokobat = form.cleaned_data['stokobat']
            jumlah = form.cleaned_data['jumlah_terima']

            so = Stokobat.objects.select_for_update().get(pk=stokobat.pk)

            obj.puskesmas = Puskesmas.objects.get(kode=self.request.user.username)
            obj.created_by = self.request.user.id
            obj.created_at = timezone.now()

            # Kurangi stok hanya di sini
            isi = int(getattr(so.masterobat, 'isi', 1)) or 1
            so.stok -= jumlah
            so.butir -= jumlah * isi
            so.save()

            # Set nilai stok dan butir di distribusi
            obj.stok = jumlah
            obj.butir = jumlah * isi
            obj.save()

            # Set ke self.object agar get_success_url() tidak error
            self.object = obj

        messages.success(self.request, "Distribusi berhasil dibuat, stok telah dikurangi.")
        return HttpResponseRedirect(self.get_success_url())

class DisObatUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    model = Distribusiobat
    template_name = 'vitamin/puskesmas/disobat/disobat_edit.html'
    form_class = DisObatForm
    success_url = reverse_lazy('vitamin:disobat-list')
    permission_required = 'vitamin.change_distribusiobat'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Ubah Distribusi Obat"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  
        return kwargs
    
    #def form_valid2(self, form):
    #    obj = form.save(commit=False)
        #puskesmas_id = UserPuskesmas(self.request)
    #    print(f' PUSKESMAS KODE {self.request.user}')
    #    user = self.request.user

        # These should ideally be based on request data
    #    obj.puskesmas = Puskesmas.objects.get(kode=user)

    #    obj.updated_by = self.request.user.id  # Make sure you have a FK to User
    #    obj.updated_at = timezone.now()

    #    obj.save()
    #    return super().form_valid(form)
    
    def form_valid(self, form):
        # Data lama (sebelum diubah)
        old_obj = self.get_object()
        old_stokobat = old_obj.stokobat
        old_qty = old_obj.jumlah_terima

        # Data baru dari form
        obj = form.save(commit=False)
        new_stokobat = form.cleaned_data['stokobat']
        new_qty = form.cleaned_data['jumlah_terima']

        # metadata update
        obj.puskesmas = Puskesmas.objects.get(kode=self.request.user.username)
        obj.updated_by = self.request.user.id
        obj.updated_at = timezone.now()

        with transaction.atomic():
            if new_stokobat.pk == old_stokobat.pk:
                # Produk sama → hitung delta
                delta = new_qty - old_qty
                if delta != 0:
                    so = Stokobat.objects.select_for_update().get(pk=new_stokobat.pk)

                    # Jika butuh stok tambahan
                    if delta > 0 and so.stok < delta:
                        form.add_error('jumlah_terima', f"Stok tidak cukup (tersedia {so.stok}).")
                        return self.form_invalid(form)

                    # Terapkan perubahan stok (delta>0 kurangi; delta<0 menambah)
                    Stokobat.objects.filter(pk=so.pk).update(stok=F('stok') - delta)
            else:
                # Produk berubah
                so_old = Stokobat.objects.select_for_update().get(pk=old_stokobat.pk)
                so_new = Stokobat.objects.select_for_update().get(pk=new_stokobat.pk)

                # Validasi stok di produk baru
                if so_new.stok < new_qty:
                    form.add_error('jumlah_terima', f"Stok tidak cukup pada item baru (tersedia {so_new.stok}).")
                    return self.form_invalid(form)

                # Kembalikan stok ke produk lama
                Stokobat.objects.filter(pk=so_old.pk).update(stok=F('stok') + old_qty)
                # Ambil stok dari produk baru
                Stokobat.objects.filter(pk=so_new.pk).update(stok=F('stok') - new_qty)

            # Simpan perubahan distribusi
            obj.save()

        messages.success(self.request, "Distribusi berhasil diubah, stok telah diperbarui.")
        return super().form_valid(form)

class DisObatDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    model = Distribusiobat
    template_name = 'vitamin/puskesmas/disobat/disobat_delete.html'
    success_url = reverse_lazy('vitamin:disobat-list')
    permission_required = 'vitamin.delete_distribusiobat'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Hapus Distribusi Obat"
        return context
    
    def form_valid(self, form):
        self.object = self.get_object()
        try:
            with transaction.atomic():
                # Ambil stok obat terkait dan kunci barisnya
                so = (
                    Stokobat.objects
                    .select_for_update()
                    .select_related("masterobat")
                    .get(pk=self.object.stokobat_id)
                )

                # Pastikan isi bertipe int
                isi = int(getattr(so.masterobat, "isi", 1) or 1)

                # Kembalikan stok dan butir
                jumlah = int(self.object.jumlah_terima)
                so.stok += jumlah
                so.butir += jumlah * isi
                so.save()

                # Hapus distribusi setelah stok dikembalikan
                response = super().form_valid(form)

            messages.success(self.request, "Distribusi berhasil dihapus. Stok & butir telah dikembalikan.")
            return response

        except ObjectDoesNotExist:
            messages.error(self.request, "Data stok tidak ditemukan. Penghapusan dibatalkan.")
            return HttpResponseRedirect(self.get_success_url())

        except ProtectedError:
            messages.error(self.request, "Distribusi tidak dapat dihapus karena dipakai di data lain.")
            return HttpResponseRedirect(self.get_success_url())

class DisObatDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    model = Distribusiobat
    template_name = 'vitamin/puskesmas/disobat/disobat_detail.html'
    success_url = reverse_lazy('vitamin:disobat-list')
    permission_required = 'vitamin.view_distribusiobat'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #mp = MenuProduk.objects.filter(produk=self.model.objects.get(id=self.kwargs.get('pk'))).first()
        #context["menu_list"] = mp.menu.all().order_by('name')
        context["title"] = "Detail Distribusi Obat"
        print(context)
        return context
    

def download_template(request):
    template_path = os.path.join(settings.BASE_DIR, 'static', 'files', 'template_distribusi_obat.xlsx')
    
    if not os.path.exists(template_path):
        raise Http404("Template file not found.")

    # Load the workbook
    wb = load_workbook(template_path)

       ### === Sheet: sekolah === ###
    sekolah_sheet = 'sekolah'
    if sekolah_sheet not in wb.sheetnames:
        raise Http404(f"Sheet '{sekolah_sheet}' not found in template.")
    ws_sekolah = wb[sekolah_sheet]

    # Clear existing sekolah sheet data (rows after header)
    for row in ws_sekolah.iter_rows(min_row=2, max_row=ws_sekolah.max_row):
        for cell in row:
            cell.value = None

    # Write sekolah data
    sekolah_data = Sekolah.objects.values_list('id', 'kode', 'nama')
    for idx, (id, kode, nama) in enumerate(sekolah_data, start=2):
        ws_sekolah.cell(row=idx, column=1).value = id
        ws_sekolah.cell(row=idx, column=2).value = kode
        ws_sekolah.cell(row=idx, column=3).value = nama
    
    ### === Sheet: vitamin === ###
    stokobat_sheet = 'stok_obat'
    if stokobat_sheet not in wb.sheetnames:
        raise Http404(f"Sheet '{stokobat_sheet}' not found in template.")
    ws_stok = wb[stokobat_sheet]

    # Clear existing vitamin sheet data (rows after header)
    for row in ws_stok.iter_rows(min_row=2, max_row=ws_stok.max_row):
        for cell in row:
            cell.value = None
    
    # Get user's associated puskesmas_id
    try:
        # Assume the username corresponds to Puskesmas.kode
        kode = request.user.username
        puskesmas = Puskesmas.objects.get(kode=kode)
        puskesmas_id = puskesmas.id
    except Puskesmas.DoesNotExist:
        messages.error(request, "Puskesmas not found for the current user.")
        return redirect('vitamin:disobat-list')

       # Filter and fetch Stokobat data for the user's puskesmas
    stokobat_data = Stokobat.objects.filter(
        puskesmas_id=puskesmas_id
    ).select_related(
        'masterobat__vitamin', 'masterobat__satuan', 'masterobat'
    ).values_list(
        'id', 'tgl_terima', 'terima', 'stok', 'keterangan',
        'masterobat__merk', 'masterobat__pabrik', 'masterobat__kadaluarsa',
        'masterobat__isi', 'masterobat__batchnumber',
        'masterobat__satuan__nama', 'masterobat__vitamin__nama'
    )

    for idx, (id, tgl_terima, terima, stok, keterangan, masterobat_merk, masterobat_pabrik, masterobat_kadaluarsa, 
              maseterobat_isi, masterobat_batchnumber, masterobat_satuan_nama, masterobat_vitamin_nama) in enumerate(stokobat_data, start=2):
        ws_stok.cell(row=idx, column=1).value = id
        ws_stok.cell(row=idx, column=2).value = tgl_terima
        ws_stok.cell(row=idx, column=3).value = terima
        ws_stok.cell(row=idx, column=4).value = stok
        ws_stok.cell(row=idx, column=5).value = keterangan
        ws_stok.cell(row=idx, column=6).value = masterobat_merk
        ws_stok.cell(row=idx, column=7).value = masterobat_pabrik
        ws_stok.cell(row=idx, column=8).value = masterobat_kadaluarsa
        ws_stok.cell(row=idx, column=9).value = maseterobat_isi
        ws_stok.cell(row=idx, column=10).value = masterobat_batchnumber
        ws_stok.cell(row=idx, column=11).value = masterobat_satuan_nama
        ws_stok.cell(row=idx, column=12).value = masterobat_vitamin_nama

    # Save workbook to memory stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return as downloadable file
    response = FileResponse(output, as_attachment=True, filename='template_distribusi_obat.xlsx')
    return response

def import_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            df = pd.read_excel(excel_file, sheet_name='distribusi_obat')

            required_columns = {
                'Tanggal Kirim', 'Tanggal Terima', 'Jumlah Terima',	'Stok',	'Id Sekolah', 'Id Stok Obat'
            }

            if not required_columns.issubset(df.columns):
                missing = required_columns - set(df.columns)
                messages.error(request, f"Missing columns in Excel file: {', '.join(missing)}")
                return redirect('vitamin:disobat-list')

            # Get puskesmas by user.username     
            try:
                puskesmas = Puskesmas.objects.get(kode=request.user.username)
            except Puskesmas.DoesNotExist:
                messages.error(request, "Puskesmas not found for the current user.")
                return redirect('vitamin:disobat-list')
            
            # ---------- PASS 1: Kumpulkan kebutuhan stok ----------
            total_needed = defaultdict(int)
            parsed_rows = []  # simpan hasil parsing untuk pass 2

            for idx, row in df.iterrows():
                if pd.isna(row['Id Stok Obat']) or pd.isna(row['Id Sekolah']):
                    raise ValueError(f"Baris {idx+2}: Id Stok Obat / Id Sekolah kosong.")

                try:
                    stokobat = Stokobat.objects.get(pk=int(row['Id Stok Obat']))
                except Stokobat.DoesNotExist:
                    raise ValueError(f"Baris {idx+2}: Stokobat id {row['Id Stok Obat']} tidak ditemukan.")

                try:
                    sekolah = Sekolah.objects.get(pk=int(row['Id Sekolah']))
                except Sekolah.DoesNotExist:
                    raise ValueError(f"Baris {idx+2}: Sekolah id {row['Id Sekolah']} tidak ditemukan.")

                qty = 0 if pd.isna(row['Jumlah Terima']) else int(row['Jumlah Terima'])
                if qty <= 0:
                    raise ValueError(f"Baris {idx+2}: Jumlah Terima tidak valid ({qty}).")

                tgl_kirim  = None if pd.isna(row['Tanggal Kirim'])  else pd.to_datetime(row['Tanggal Kirim'])
                tgl_terima = None if pd.isna(row['Tanggal Terima']) else pd.to_datetime(row['Tanggal Terima'])

                total_needed[stokobat.pk] += qty
                parsed_rows.append((stokobat, sekolah, tgl_kirim, tgl_terima, qty))

            # ---------- PASS 2: Validasi stok & buat data ----------
            with transaction.atomic():
                # Lock semua stok yang terlibat
                stok_map = {
                    so.pk: so
                    for so in Stokobat.objects.select_for_update().filter(pk__in=total_needed.keys())
                }

                # Cek ketersediaan
                for stok_id, needed_qty in total_needed.items():
                    if stok_id not in stok_map:
                        raise ValueError(f"Stok id {stok_id} tidak ditemukan.")
                    if stok_map[stok_id].stok < needed_qty:
                        raise ValueError(
                            f"Stok {stok_map[stok_id]} tidak cukup "
                            f"(tersedia {stok_map[stok_id].stok}, dibutuhkan {needed_qty})."
                        )

                # Jika semua cukup, lakukan insert dan update stok
                for stokobat, sekolah, tgl_kirim, tgl_terima, qty in parsed_rows:
                    Distribusiobat.objects.create(
                        tgl_kirim=tgl_kirim,
                        tgl_terima=tgl_terima,
                        jumlah_terima=qty,
                        stok=stok_map[stokobat.pk].stok,  # optional catatan stok saat ini
                        puskesmas=puskesmas,
                        sekolah=sekolah,
                        stokobat=stokobat,
                        created_by=request.user.id,
                        created_at=timezone.now(),
                    )
                    Stokobat.objects.filter(pk=stokobat.pk).update(stok=F('stok') - qty)

            messages.success(request, f"Import selesai: {len(parsed_rows)} baris dimasukkan. Stok diperbarui.")

        except ValueError as ve:
            messages.error(request, f"Import dibatalkan: {ve}")

        except Exception as e:
            messages.error(request, f"Error importing file: {e}")

    return redirect('vitamin:disobat-list')