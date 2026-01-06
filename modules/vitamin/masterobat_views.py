from django.shortcuts import render, redirect
from django.views.generic import ListView,DetailView, TemplateView
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.http import request,HttpResponseRedirect
from django.db.models import Q

from modules.vitamin.models import MasterObat

from modules.core.core_libs import *
from django.contrib.auth.models import Group

from django.contrib.auth.mixins import PermissionRequiredMixin
import logging

logger = logging.getLogger(__name__)
from django.forms import formset_factory
from modules.uman.decorators  import group_must_have_permission
from django.utils.decorators import method_decorator
from django.views import View
from datetime import date
from django.core.exceptions import ObjectDoesNotExist

from .forms import *

import os
from django.conf import settings
from django.http import FileResponse, Http404
from openpyxl import load_workbook
from io import BytesIO

import pandas as pd
from django.shortcuts import redirect
from django.contrib import messages
from .models import Vitamin, Satuan  # Adjust model names to match your actual models



# ----------------referensi Vitamin--------------------

@method_decorator([group_must_have_permission], name='dispatch')
class MasterObatListView(LoginRequiredMixin,PermissionRequiredMixin, ListView):
    model = MasterObat
    #paginate_by = 10
    permission_required = 'vitamin.view_masterobat'
    def get_template_names(self):
        if self.request.htmx:
            return ["vitamin/masterobat/masterobat_list.html"] # The response HTML to inject into a list
        else:
            return ["vitamin/masterobat/masterobat.html"] # The actual form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["numlist"] = (int(self.request.GET.get("page",1)) - 1) 
        context["q"] = self.request.GET.get("q",'')
        context["title"] = "Daftar Master Obat"
        return context
    
    def get_queryset(self):
        qs = self.model.objects.all()
        if self.request.GET.get("q",''):
              qs = self.model.objects.filter(
                 Q(merk__icontains=self.request.GET.get("q",'')) | 
                 Q(pabrik__icontains=self.request.GET.get("q",'')) | 
                 Q(batchnumber__icontains=self.request.GET.get("q",'')) | 
                 Q(vitamin__nama__icontains=self.request.GET.get("q",'')) | 
                 Q(satuan__nama__icontains=self.request.GET.get("q",''))
              )
        return qs.order_by('-id')
    

class MasterObatCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    model = MasterObat
    template_name = 'vitamin/masterobat/masterobat_form.html'
    form_class = MasterObatForm
    success_url = reverse_lazy('vitamin:masterobat-list')
    permission_required = 'vitamin.add_masterobat'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Tambah Master Obat"
        return context
    

class MasterObatUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    model = MasterObat
    template_name = 'vitamin/masterobat/masterobat_edit.html'
    form_class = MasterObatForm
    success_url = reverse_lazy('vitamin:masterobat-list')
    permission_required = 'vitamin.change_masterobat'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Ubah Master Obat"
        return context

class MasterObatDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    model = MasterObat
    template_name = 'vitamin/masterobat/masterobat_delete.html'
    success_url = reverse_lazy('vitamin:masterobat-list')
    permission_required = 'vitamin.delete_masterobat'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Hapus Vitamin"
        return context

class MasterObatDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    model = MasterObat
    template_name = 'vitamin/masterobat/masterobat_detail.html'
    success_url = reverse_lazy('vitamin:masterobat-list')
    permission_required = 'vitamin.view_masterobat'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #mp = MenuProduk.objects.filter(produk=self.model.objects.get(id=self.kwargs.get('pk'))).first()
        #context["menu_list"] = mp.menu.all().order_by('name')
        context["title"] = "Detail Master Obat"
        print(context)
        return context
    


def download_template(request):
    template_path = os.path.join(settings.BASE_DIR, 'static', 'files', 'template_master_obat.xlsx')
    
    if not os.path.exists(template_path):
        raise Http404("Template file not found.")

    # Load the workbook
    wb = load_workbook(template_path)
    
    ### === Sheet: vitamin === ###
    vitamin_sheet = 'vitamin'
    if vitamin_sheet not in wb.sheetnames:
        raise Http404(f"Sheet '{vitamin_sheet}' not found in template.")
    ws_vitamin = wb[vitamin_sheet]

    # Clear existing vitamin sheet data (rows after header)
    for row in ws_vitamin.iter_rows(min_row=2, max_row=ws_vitamin.max_row):
        for cell in row:
            cell.value = None

    # Write Vitamin data
    vitamin_data = Vitamin.objects.values_list('id', 'nama')
    for idx, (id, nama) in enumerate(vitamin_data, start=2):
        ws_vitamin.cell(row=idx, column=1).value = id
        ws_vitamin.cell(row=idx, column=2).value = nama

    ### === Sheet: satuan === ###
    satuan_sheet = 'satuan'
    if satuan_sheet not in wb.sheetnames:
        raise Http404(f"Sheet '{satuan_sheet}' not found in template.")
    ws_satuan = wb[satuan_sheet]

    # Clear existing satuan sheet data (rows after header)
    for row in ws_satuan.iter_rows(min_row=2, max_row=ws_satuan.max_row):
        for cell in row:
            cell.value = None

    # Write Satuan data
    satuan_data = Satuan.objects.values_list('id', 'nama', 'keterangan')
    for idx, (id, nama, keterangan) in enumerate(satuan_data, start=2):
        ws_satuan.cell(row=idx, column=1).value = id
        ws_satuan.cell(row=idx, column=2).value = nama
        ws_satuan.cell(row=idx, column=3).value = keterangan

    # Save workbook to memory stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return as downloadable file
    response = FileResponse(output, as_attachment=True, filename='template_master_obat.xlsx')
    return response


def import_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            # Load Excel sheet
            df = pd.read_excel(excel_file, sheet_name='master_obat')

            # Define required columns
            required_columns = {
                'Merk', 'Pabrik', 'Tanggal Kadaluarsa',
                'Isi', 'Batchnumber', 'Id Vitamin', 'Id Satuan'
            }

            # Check for missing columns
            if not required_columns.issubset(df.columns):
                missing = required_columns - set(df.columns)
                messages.error(request, f"Missing columns in Excel file: {', '.join(missing)}")
                return redirect('vitamin:satuan-list')

            # Iterate over DataFrame rows
            for _, row in df.iterrows():
                try:
                    # Skip rows missing critical fields
                    if pd.isna(row['Id Vitamin']) or pd.isna(row['Id Satuan']):
                        continue

                    vitamin = Vitamin.objects.get(pk=int(row['Id Vitamin']))
                    satuan = Satuan.objects.get(pk=int(row['Id Satuan']))

                    MasterObat.objects.create(
                        merk=str(row['Merk']),
                        pabrik=str(row['Pabrik']) if not pd.isna(row['Pabrik']) else '',
                        kadaluarsa=pd.to_datetime(row['Tanggal Kadaluarsa']) if not pd.isna(row['Tanggal Kadaluarsa']) else None,
                        isi=row['Isi'],
                        batchnumber=str(row['Batchnumber']) if not pd.isna(row['Batchnumber']) else '',
                        vitamin=vitamin,
                        satuan=satuan
                    )
                except Vitamin.DoesNotExist:
                    messages.warning(request, f"Vitamin with ID {row['Id Vitamin']} not found.")
                except Satuan.DoesNotExist:
                    messages.warning(request, f"Satuan with ID {row['Id Satuan']} not found.")
                except Exception as row_err:
                    messages.warning(request, f"Error in row: {row_err}")

            messages.success(request, "Excel file imported successfully.")
        except Exception as e:
            messages.error(request, f"Error importing file: {e}")

    return redirect('vitamin:masterobat-list')
