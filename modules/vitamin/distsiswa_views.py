from django.shortcuts import render, redirect
from django.views.generic import ListView,DetailView, TemplateView
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.http import request,HttpResponseRedirect
from django.db.models import Q, Sum, Min, Max
from modules.vitamin.models import Distribusisiswa

from modules.core.core_libs import *
from django.contrib.auth.models import Group

from django.contrib.auth.mixins import PermissionRequiredMixin
import logging

logger = logging.getLogger(__name__)
from django.forms import formset_factory
from modules.uman.decorators  import group_must_have_permission
from django.utils.decorators import method_decorator
from django.views import View
from datetime import date
from django.core.exceptions import ObjectDoesNotExist

from .forms import *

from django.utils import timezone

import os
from django.conf import settings
from django.http import FileResponse, Http404
from openpyxl import load_workbook
from io import BytesIO

import pandas as pd
from django.shortcuts import redirect
from django.contrib import messages
from .models import Vitamin, Satuan, Stokobat  # Adjust model names to match your actual models
from django.db.models import F
from django.db.models.deletion import ProtectedError
from collections import defaultdict
# ----------------referensi distsiswa--------------------

@method_decorator([group_must_have_permission], name='dispatch')
class DistSiswaListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Distribusisiswa
    permission_required = 'vitamin.view_distribusisiswa'

    def get_template_names(self):
        if self.request.htmx:
            return ["vitamin/sekolah/distsiswa/distsiswa_list.html"]  # untuk partial HTMX
        else:
            return ["vitamin/sekolah/distsiswa/distsiswa.html"]  # untuk halaman utama

    def get_queryset(self):
        search_query = self.request.GET.get("q", "").strip()
        app_group = Group.objects.get(name='administrator')

        base_qs = self.model.objects.all()

        if not self.request.user.is_superuser and app_group not in self.request.user.groups.all():
            try:
                sekolah_id = self.request.user.sekolah.id
            except Exception:
                sekolah_id = 0
            base_qs = base_qs.filter(sekolah_id=sekolah_id)

        # ✅ Tetap return instance model, bukan dict
        qs = (
    base_qs
    .values(
        "nis",
        "nama_siswa",
        "kelas",
        "sekolah__nama",
        "vitamin__nama",
    )
    .annotate(
        total_jumlah=Sum("jumlah"),
        first_id=Min("id"),   # ✅ ambil salah satu id untuk url aksi
        last_tgl=Max("tgl_terima"),  # ✅ ambil tanggal terakhir terima
    )
    .order_by("-last_tgl", "-nis")
)


        if search_query:
            qs = qs.filter(
                Q(nis__icontains=search_query) |
                Q(nama_siswa__icontains=search_query) |
                Q(kelas__icontains=search_query) |
                Q(sekolah__nama__icontains=search_query) |
                Q(vitamin__nama__icontains=search_query)
            )

        return qs
class DistSiswaCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    model = Distribusisiswa
    template_name = 'vitamin/sekolah/distsiswa/distsiswa_form.html'
    form_class = DistSiswaForm
    success_url = reverse_lazy('vitamin:distsiswa-list')
    permission_required = 'vitamin.add_distribusisiswa'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Tambah Distribusi Siswa"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # ✅ Pass the logged-in user to the form
        return kwargs
    
    def form_valid(self, form):
        obj = form.save(commit=False)

                # Expect these fields to exist in the form:
        # - stokobat  (FK to StokObat)
        # - jumlah_terima (int)
        obj = form.save(commit=False)
        distribusiobat = form.cleaned_data['distribusiobat']
        jumlah = form.cleaned_data['jumlah']

        with transaction.atomic():
                # Lock the stock row so concurrent requests can't oversell
                so = Distribusiobat.objects.select_for_update().get(pk=distribusiobat.pk)

                user = self.request.user

                # These should ideally be based on request data
                obj.sekolah = Sekolah.objects.get(kode=user)

                # Ambil siswa yang dipilih dari form
                siswa = form.cleaned_data.get('siswa')
                if not siswa:
                    form.add_error('siswa', "Siswa wajib dipilih.")
                    return self.form_invalid(form)

                # Isi otomatis dari siswa
                obj.nis = getattr(siswa, 'nis', None)       # pastikan field di Siswa bernama 'nis'
                obj.nama_siswa = getattr(siswa, 'nama', None)     # atau 'nama_lengkap' sesuai model kamu

                obj.created_by = self.request.user.id  # Make sure you have a FK to User
                obj.created_at = timezone.now() 
                            
                obj.save()

                # Decrement stock safely in the DB
                Distribusiobat.objects.filter(pk=so.pk).update(stok=F('stok') - jumlah)

        messages.success(self.request, "Distribusi Siswa berhasil dibuat, stok telah dikurangi.")
        return super().form_valid(form)

        

class DistSiswaUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    model = Distribusisiswa
    template_name = 'vitamin/sekolah/distsiswa/distsiswa_edit.html'
    form_class = DistSiswaForm
    success_url = reverse_lazy('vitamin:distsiswa-list')
    permission_required = 'vitamin.change_distribusisiswa'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Ubah Distribusi Siswa"
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # ✅ Pass the logged-in user to the form
        return kwargs
    
    #def form_valid(self, form):
    #    obj = form.save(commit=False)

    #    user = self.request.user

        # These should ideally be based on request data
    #    obj.sekolah = Sekolah.objects.get(kode=user)

    #    obj.updated_by = self.request.user.id  # Make sure you have a FK to User
    #    obj.updated_at = timezone.now()

    #    obj.save()
    #    return super().form_valid(form)
    
    def form_valid(self, form):
        # Data lama (sebelum diubah)
        old_obj = self.get_object()
        old_distribusiobat = old_obj.distribusiobat
        old_qty = old_obj.jumlah

        # Data baru dari form
        obj = form.save(commit=False)
        new_distribusiobat = form.cleaned_data['distribusiobat']
        new_qty = form.cleaned_data['jumlah']

        # metadata update
        user = self.request.user

        # These should ideally be based on request data
        obj.sekolah = Sekolah.objects.get(kode=user)
        obj.updated_by = self.request.user.id
        obj.updated_at = timezone.now()

        with transaction.atomic():
            if new_distribusiobat.pk == old_distribusiobat.pk:
                # Produk sama → hitung delta
                delta = new_qty - old_qty
                if delta != 0:
                    so = Distribusiobat.objects.select_for_update().get(pk=new_distribusiobat.pk)

                    # Jika butuh stok tambahan
                    if delta > 0 and so.stok < delta:
                        form.add_error('jumlah', f"Stok tidak cukup (tersedia {so.stok}).")
                        return self.form_invalid(form)

                    # Terapkan perubahan stok (delta>0 kurangi; delta<0 menambah)
                    Distribusiobat.objects.filter(pk=so.pk).update(stok=F('stok') - delta)
            else:
                # Produk berubah
                so_old = Distribusiobat.objects.select_for_update().get(pk=old_distribusiobat.pk)
                so_new = Distribusiobat.objects.select_for_update().get(pk=new_distribusiobat.pk)

                # Validasi stok di produk baru
                if so_new.stok < new_qty:
                    form.add_error('jumlah', f"Stok tidak cukup pada item baru (tersedia {so_new.stok}).")
                    return self.form_invalid(form)

                # Kembalikan stok ke produk lama
                Distribusiobat.objects.filter(pk=so_old.pk).update(stok=F('stok') + old_qty)
                # Ambil stok dari produk baru
                Distribusiobat.objects.filter(pk=so_new.pk).update(stok=F('stok') - new_qty)

            # Simpan perubahan distribusi
            obj.save()

        messages.success(self.request, "Distribusi Siswa berhasil diubah, stok telah diperbarui.")
        return super().form_valid(form)

class DistSiswaDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    model = Distribusisiswa
    template_name = 'vitamin/sekolah/distsiswa/distsiswa_delete.html'
    success_url = reverse_lazy('vitamin:distsiswa-list')
    permission_required = 'vitamin.delete_distribusisiswa'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Hapus Distribusi Siswa"
        return context
    
    def form_valid(self, form):
        # Django akan memanggil ini saat POST konfirmasi hapus
        self.object = self.get_object()
        try:
            with transaction.atomic():
                # Kunci baris stok agar konsisten
                so = Distribusiobat.objects.select_for_update().get(pk=self.object.distribusiobat_id)

                # Kembalikan stok sesuai jumlah_terima
                Distribusiobat.objects.filter(pk=so.pk).update(
                    stok=F('stok') + self.object.jumlah
                )

                # Lanjutkan ke default delete milik DeleteView
                response = super().form_valid(form)

            messages.success(self.request, "Distribusi berhasil dihapus. Stok telah dikembalikan.")
            return response

        except ObjectDoesNotExist:
            messages.error(self.request, "Data stok tidak ditemukan. Penghapusan dibatalkan.")
            return HttpResponseRedirect(self.get_success_url())

        except ProtectedError:
            # Rollback otomatis karena atomic
            messages.error(self.request, "Distribusi tidak dapat dihapus karena dipakai di data lain.")
            return HttpResponseRedirect(self.get_success_url())

class DistSiswaDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    model = Distribusisiswa
    template_name = 'vitamin/sekolah/distsiswa/distsiswa_detail.html'
    success_url = reverse_lazy('vitamin:distsiswa-list')
    permission_required = 'vitamin.view_distribusisiswa'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #mp = MenuProduk.objects.filter(produk=self.model.objects.get(id=self.kwargs.get('pk'))).first()
        #context["menu_list"] = mp.menu.all().order_by('name')
        context["title"] = "Detail Distribusi Siswa"
        print(context)
        return context
    

def download_template(request):
    template_path = os.path.join(settings.BASE_DIR, 'static', 'files', 'template_distribusi_siswa.xlsx')
    
    if not os.path.exists(template_path):
        raise Http404("Template file not found.")

    # Load the workbook
    wb = load_workbook(template_path)

       ### === Sheet: vitamin === ###
    vitamin_sheet = 'vitamin'
    if vitamin_sheet not in wb.sheetnames:
        raise Http404(f"Sheet '{vitamin_sheet}' not found in template.")
    ws_vitamin = wb[vitamin_sheet]

    # Clear existing vitamin sheet data (rows after header)
    for row in ws_vitamin.iter_rows(min_row=2, max_row=ws_vitamin.max_row):
        for cell in row:
            cell.value = None

    # Write vitamin data
    vitamin_data = Vitamin.objects.values_list('id', 'nama')
    for idx, (id, nama) in enumerate(vitamin_data, start=2):
        ws_vitamin.cell(row=idx, column=1).value = id
        ws_vitamin.cell(row=idx, column=2).value = nama


    ### === Sheet: Siswa === ###
    siswa_sheet = 'siswa'
    if siswa_sheet not in wb.sheetnames:
        raise Http404(f"Sheet '{siswa_sheet}' not found in template.")
    ws_siswa = wb[siswa_sheet]

    # Clear existing vitamin sheet data (rows after header)
    for row in ws_siswa.iter_rows(min_row=2, max_row=ws_siswa.max_row):
        for cell in row:
            cell.value = None
 # Get user's associated puskesmas_id
    try:
        # Assume the username corresponds to Puskesmas.kode
        kode = request.user.username
        sekolah = Sekolah.objects.get(kode=kode)
        sekolah_id = sekolah.id
    except Sekolah.DoesNotExist:
        messages.error(request, "Sekolah not found for the current user.")
        return redirect('vitamin:distsiswa-list')

       # Filter and fetch Stokobat data for the user's puskesmas
    siswa_data = Siswa.objects.filter(
        sekolah_id=sekolah_id
    ).values_list(
        'id', 'nis', 'nama'
    )

    for idx, (id, nis, nama) in enumerate(siswa_data, start=2):
        ws_siswa.cell(row=idx, column=1).value = id
        ws_siswa.cell(row=idx, column=2).value = nis
        ws_siswa.cell(row=idx, column=3).value = nama
       
     ### === Sheet: Distribusi Obat === ###
    distribusiobat_sheet = 'distribusi_obat'
    if distribusiobat_sheet not in wb.sheetnames:
        raise Http404(f"Sheet '{distribusiobat_sheet}' not found in template.")
    ws_distobat = wb[distribusiobat_sheet]

    # Clear existing distribusiobat sheet data (rows after header)
    for row in ws_distobat.iter_rows(min_row=2, max_row=ws_distobat.max_row):
        for cell in row:
            cell.value = None
    
    try:
        # Assume the username corresponds to Puskesmas.kode
        kode = request.user.username
        sekolah = Sekolah.objects.get(kode=kode)
        sekolah_id = sekolah.id
    except Sekolah.DoesNotExist:
        messages.error(request, "Sekolah not found for the current user.")
        return redirect('vitamin:distsiswa-list')

       # Filter and fetch Stokobat data for the user's puskesmas
    distribusiobat_data = Distribusiobat.objects.filter(
        sekolah_id=sekolah_id
    ).select_related(
        'stokobat__masterobat__vitamin', 'vitamin', 'stokobat','masterobat'
    ).values_list(
        'id', 'tgl_kirim', 'tgl_terima', 'jumlah_terima', 'stok',
        'stokobat__masterobat__vitamin__nama'
    )

    for idx, (id, tgl_kirim, tgl_terima, jumlah_terima, stok, stokobat_masterobat_vitamin_nama) in enumerate(distribusiobat_data, start=2):
        ws_distobat.cell(row=idx, column=1).value = id
        ws_distobat.cell(row=idx, column=2).value = tgl_kirim
        ws_distobat.cell(row=idx, column=3).value = tgl_terima
        ws_distobat.cell(row=idx, column=4).value = jumlah_terima
        ws_distobat.cell(row=idx, column=5).value = stok
        ws_distobat.cell(row=idx, column=6).value = stokobat_masterobat_vitamin_nama

    # Save workbook to memory stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return as downloadable file
    response = FileResponse(output, as_attachment=True, filename='template_distribusi_siswa.xlsx')
    return response


from collections import defaultdict
from django.db import transaction
from django.db.models import F
from django.contrib import messages
from django.utils import timezone
import pandas as pd

def import_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            df = pd.read_excel(excel_file, sheet_name='distribusi_obat_siswa')

            required_columns = {
                'Id Siswa', 'Jumlah Terima', 'Tanggal Terima',
                'Kelas', 'Id Vitamin', 'Id Distribusi Obat'
            }
            if not required_columns.issubset(df.columns):
                missing = required_columns - set(df.columns)
                messages.error(request, f"Missing columns in Excel file: {', '.join(missing)}")
                return redirect('vitamin:distsiswa-list')

            # Sekolah user (opsional: bisa pakai siswa.sekolah agar 100% konsisten)
            try:
                sekolah_user = Sekolah.objects.get(kode=request.user.username)
            except Sekolah.DoesNotExist:
                messages.error(request, "Sekolah not found for the current user.")
                return redirect('vitamin:distsiswa-list')

            # ---------- PASS 1: Parse & agregasi kebutuhan stok per Distribusiobat ----------
            total_needed = defaultdict(int)
            parsed_rows = []  # simpan data siap pakai untuk PASS 2

            for idx, row in df.iterrows():
                rownum = idx + 2  # baris excel (header = baris 1)

                # Validasi id-id wajib
                if pd.isna(row['Id Siswa']) or pd.isna(row['Id Vitamin']) or pd.isna(row['Id Distribusi Obat']):
                    raise ValueError(f"Baris {rownum}: Id Siswa / Id Vitamin / Id Distribusi Obat kosong.")

                try:
                    siswa = Siswa.objects.get(pk=int(row['Id Siswa']))
                except Siswa.DoesNotExist:
                    raise ValueError(f"Baris {rownum}: Siswa id {row['Id Siswa']} tidak ditemukan.")

                try:
                    vitamin = Vitamin.objects.get(pk=int(row['Id Vitamin']))
                except Vitamin.DoesNotExist:
                    raise ValueError(f"Baris {rownum}: Vitamin id {row['Id Vitamin']} tidak ditemukan.")

                try:
                    dist = Distribusiobat.objects.get(pk=int(row['Id Distribusi Obat']))
                except Distribusiobat.DoesNotExist:
                    raise ValueError(f"Baris {rownum}: Distribusiobat id {row['Id Distribusi Obat']} tidak ditemukan.")

                # Qty & fields lain
                qty = 0 if pd.isna(row['Jumlah Terima']) else int(row['Jumlah Terima'])
                if qty <= 0:
                    raise ValueError(f"Baris {rownum}: Jumlah Terima tidak valid ({qty}).")

                tgl_terima = None if pd.isna(row['Tanggal Terima']) else pd.to_datetime(row['Tanggal Terima'])
                kelas = (row['Kelas'] if not pd.isna(row['Kelas']) else "") or ""

                # (Opsional) pastikan siswa sesuai sekolah user
                # if siswa.sekolah_id != sekolah_user.id:
                #     raise ValueError(f"Baris {rownum}: Siswa bukan dari sekolah Anda.")

                parsed_rows.append({
                    "rownum": rownum,
                    "siswa": siswa,
                    "vitamin": vitamin,
                    "dist": dist,
                    "qty": qty,
                    "tgl_terima": tgl_terima,
                    "kelas": kelas,
                })
                total_needed[dist.pk] += qty

            # ---------- PASS 2: Validasi stok & eksekusi (satu transaksi) ----------
            with transaction.atomic():
                # Kunci semua distribusi yang akan dipakai agar aman dari race condition
                locked_dists = {
                    d.pk: d
                    for d in Distribusiobat.objects.select_for_update().filter(pk__in=total_needed.keys())
                }

                # Cek sisa stok per Distribusiobat
                for dist_id, needed in total_needed.items():
                    dist = locked_dists.get(dist_id)
                    if not dist:
                        raise ValueError(f"Distribusiobat id {dist_id} tidak ditemukan saat validasi stok.")
                    if dist.stok < needed:
                        raise ValueError(
                            f"Sisa stok untuk Distribusiobat #{dist_id} tidak cukup "
                            f"(tersedia {dist.stok}, dibutuhkan {needed})."
                        )

                # Semua cukup → buat Distribusisiswa dan kurangi stok parent
                for item in parsed_rows:
                    s = item["siswa"]
                    v = item["vitamin"]
                    dist = item["dist"]
                    qty = item["qty"]

                    Distribusisiswa.objects.create(
                        nis=s.nis,
                        nama_siswa=s.nama,
                        siswa=siswa,
                        jumlah=qty,
                        tgl_terima=item["tgl_terima"],
                        kelas=item["kelas"] or getattr(s, 'kelas', ''),
                        sekolah=sekolah_user,         # atau pakai s.sekolah bila ingin 100% ikut siswa
                        vitamin=v,
                        distribusiobat=dist,
                        created_by=request.user.id,      # jika FK User; kalau IntegerField gunakan .id
                        created_at=timezone.now(),
                    )

                    # Kurangi sisa stok pada Distribusiobat ini
                    Distribusiobat.objects.filter(pk=dist.pk).update(stok=F('stok') - qty)

            messages.success(request, f"Import selesai: {len(parsed_rows)} baris dimasukkan. Stok diperbarui.")

        except ValueError as ve:
            messages.error(request, f"Import dibatalkan: {ve}")
        except Exception as e:
            messages.error(request, f"Error importing file: {e}")

    return redirect('vitamin:distsiswa-list')
