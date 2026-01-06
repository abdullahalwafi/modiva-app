from django.shortcuts import render, redirect
from django.views.generic import ListView,DetailView, TemplateView
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.http import request,HttpResponseRedirect
from django.db.models import Q

from modules.vitamin.models import Stokobat

from modules.core.core_libs import *
from django.contrib.auth.models import Group

from django.contrib.auth.mixins import PermissionRequiredMixin
import logging

logger = logging.getLogger(__name__)
from django.forms import formset_factory
from modules.uman.decorators  import group_must_have_permission
from django.utils.decorators import method_decorator
from django.views import View
from datetime import date
from django.core.exceptions import ObjectDoesNotExist

from .forms import *

from django.utils import timezone

import os
from django.conf import settings
from django.http import FileResponse, Http404
from openpyxl import load_workbook
from io import BytesIO

import pandas as pd
from django.shortcuts import redirect
from django.contrib import messages
from .models import Vitamin, Satuan, Stokobat  # Adjust model names to match your actual models
from django.http import JsonResponse

# ----------------referensi StokObat--------------------

@method_decorator([group_must_have_permission], name='dispatch')
class StokObatListView(LoginRequiredMixin,PermissionRequiredMixin, ListView):
    model = Stokobat
    #paginate_by = 10
    permission_required = 'vitamin.view_stokobat'
    def get_template_names(self):
        if self.request.htmx:
            return ["vitamin/puskesmas/stokobat/stokobat_list.html"] # The response HTML to inject into a list
        else:
            return ["vitamin/puskesmas/stokobat/stokobat.html"] # The actual form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["numlist"] = (int(self.request.GET.get("page",1)) - 1) 
        context["q"] = self.request.GET.get("q",'')
        context["title"] = "Daftar Stok Obat"
        
        puskesmas_id = self.request.GET.get('id', '0')

        #-------------awal jika user admin-------------
        app_group = Group.objects.get(name='administrator')

        if not self.request.user.is_superuser and app_group not in self.request.user.groups.all():
            try:
                puskesmas_id = Puskesmas.objects.get(kode=self.request.user).id
            except Exception as e:
                puskesmas_id = 0
            qs = self.model.objects.filter(puskesmas_id=puskesmas_id).count()

        else:
            qs = self.model.objects.filter().count() 
        
        context["count"] = qs

        return context
    
    
    def get_queryset(self):
        
        user_id = self.request.user.id

        search_query = self.request.GET.get("q", "").strip()

        #-------------awal jika user admin-------------
        app_group = Group.objects.get(name='administrator')

        print(self.request.user.is_superuser)
        print(self.request.user.groups.all())

        if not self.request.user.is_superuser and app_group not in self.request.user.groups.all():
            try:
                print(user_id)
                print(self.request.user)
                puskesmas_id = Puskesmas.objects.get(kode=self.request.user).id
                print(puskesmas_id)
                print('AAAAAAAAAAA')
            except Exception as e:
                puskesmas_id = 0
                print('BBBBBBBBBB')
            qs = self.model.objects.filter(puskesmas_id=puskesmas_id).order_by('-tgl_terima','-id')

            if search_query:
                qs = qs.filter(
                    Q(keterangan__icontains=search_query) |
                    Q(tgl_terima__icontains=search_query)
                )
            return qs

        else:
            qs = self.model.objects.filter().order_by('-tgl_terima','-id')
            print('AAAAAAAAAAA2222222222')
            if search_query:
                qs = qs.filter(
                    Q(puskesmas__nama__icontains=search_query) |
                    Q(tgl_terima__icontains=search_query)
                )

            return qs
        # -------------akhir jika user admin------
    

class StokObatCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Stokobat
    template_name = 'vitamin/puskesmas/stokobat/stokobat_form.html'
    form_class = StokObatForm
    success_url = reverse_lazy('vitamin:stokobat-list')
    permission_required = 'vitamin.add_stokobat'

    def form_valid(self, form):
        obj = form.save(commit=False)

        user = self.request.user
        obj.puskesmas = Puskesmas.objects.get(kode=user)
        obj.created_by = user.id
        obj.created_at = timezone.now()

        # stok otomatis = jumlah diterima
        obj.stok = obj.terima

        # hitung butir otomatis = terima * isi (dari master obat)
        try:
            isi = int(obj.masterobat.isi or 0)
            obj.butir = (obj.terima or 0) * isi
        except (ValueError, TypeError):
            obj.butir = None

        obj.save()
        return super().form_valid(form)


def get_masterobat_isi(request):
    masterobat_id = request.GET.get('id')
    try:
        masterobat = MasterObat.objects.get(id=masterobat_id)
        isi = int(masterobat.isi or 0)
    except (MasterObat.DoesNotExist, ValueError, TypeError):
        isi = 0
    return JsonResponse({'isi': isi})


class StokObatUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Stokobat
    template_name = 'vitamin/puskesmas/stokobat/stokobat_edit.html'
    form_class = StokObatForm
    success_url = reverse_lazy('vitamin:stokobat-list')
    permission_required = 'vitamin.change_stokobat'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Ubah Stok Obat"
        return context
    
    def form_valid(self, form):
        obj = form.save(commit=False)
        user = self.request.user

        obj.puskesmas = Puskesmas.objects.get(kode=user)

        obj.updated_by = self.request.user.id  
        obj.updated_at = timezone.now()

        # hitung ulang butir = stok * isi
        try:
            isi = int(obj.masterobat.isi or 0)
            obj.butir = (obj.stok or 0) * isi
        except (ValueError, TypeError):
            obj.butir = None

        obj.save()
        return super().form_valid(form)


class StokObatDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    model = Stokobat
    template_name = 'vitamin/puskesmas/stokobat/stokobat_delete.html'
    success_url = reverse_lazy('vitamin:stokobat-list')
    permission_required = 'vitamin.delete_stokobat'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Hapus Stok Obat"
        return context

class StokObatDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    model = Stokobat
    template_name = 'vitamin/puskesmas/stokobat/stokobat_detail.html'
    success_url = reverse_lazy('vitamin:stokobat-list')
    permission_required = 'vitamin.view_stokobat'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #mp = MenuProduk.objects.filter(produk=self.model.objects.get(id=self.kwargs.get('pk'))).first()
        #context["menu_list"] = mp.menu.all().order_by('name')
        context["title"] = "Detail Stok Obat"
        print(context)
        return context
    

def download_template(request):
    template_path = os.path.join(settings.BASE_DIR, 'static', 'files', 'template_stok_obat.xlsx')
    
    if not os.path.exists(template_path):
        raise Http404("Template file not found.")

    # Load the workbook
    wb = load_workbook(template_path)
    
    ### === Sheet: vitamin === ###
    masterobat_sheet = 'master_obat'
    if masterobat_sheet not in wb.sheetnames:
        raise Http404(f"Sheet '{masterobat_sheet}' not found in template.")
    ws_masterobat = wb[masterobat_sheet]

    # Clear existing vitamin sheet data (rows after header)
    for row in ws_masterobat.iter_rows(min_row=2, max_row=ws_masterobat.max_row):
        for cell in row:
            cell.value = None

    # Write Vitamin data
    masterobat_data = MasterObat.objects.select_related('satuan', 'vitamin').values_list(
    'id', 'merk', 'pabrik', 'kadaluarsa', 'isi', 'batchnumber', 'satuan__nama', 'vitamin__nama')

    for idx, (id, merk, pabrik, kadaluarsa, isi, batchnumber, satuan_nama, vitamin_nama) in enumerate(masterobat_data, start=2):
        ws_masterobat.cell(row=idx, column=1).value = id
        ws_masterobat.cell(row=idx, column=2).value = merk
        ws_masterobat.cell(row=idx, column=3).value = pabrik
        ws_masterobat.cell(row=idx, column=4).value = kadaluarsa
        ws_masterobat.cell(row=idx, column=5).value = isi
        ws_masterobat.cell(row=idx, column=6).value = batchnumber
        ws_masterobat.cell(row=idx, column=7).value = satuan_nama
        ws_masterobat.cell(row=idx, column=8).value = vitamin_nama

    # Save workbook to memory stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return as downloadable file
    response = FileResponse(output, as_attachment=True, filename='template_stok_obat.xlsx')
    return response

def import_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            df = pd.read_excel(excel_file, sheet_name='stok_obat')

            required_columns = {
                'Tanggal Terima', 'Jumlah Terima', 'Stok', 'Keterangan', 'Id Master Obat'
            }

            if not required_columns.issubset(df.columns):
                missing = required_columns - set(df.columns)
                messages.error(request, f"Missing columns in Excel file: {', '.join(missing)}")
                return redirect('vitamin:stokobat-list')

            # Get puskesmas by user.username
            try:
                puskesmas = Puskesmas.objects.get(kode=request.user.username)
            except Puskesmas.DoesNotExist:
                messages.error(request, "Puskesmas not found for the current user.")
                return redirect('vitamin:stokobat-list')

            for _, row in df.iterrows():
                try:

                    if pd.isna(row['Id Master Obat']):
                        continue

                    masterobat = MasterObat.objects.get(pk=int(row['Id Master Obat']))
                    
                    # Create related StokObat entry
                    Stokobat.objects.create(
                        tgl_terima=pd.to_datetime(row['Tanggal Terima']) if not pd.isna(row['Tanggal Terima']) else None,
                        terima=int(row['Jumlah Terima']) if not pd.isna(row['Jumlah Terima']) else 0,
                        stok=int(row['Stok']) if not pd.isna(row['Stok']) else 0,
                        keterangan=str(row['Keterangan']) if not pd.isna(row['Keterangan']) else '',
                        puskesmas=puskesmas,
                        masterobat=masterobat
                    )

                except MasterObat.DoesNotExist:
                    messages.warning(request, f"MasterObat with ID {row['Id Master Obat']} not found.")
                except Exception as row_err:
                    messages.warning(request, f"Error in row: {row_err}")

            messages.success(request, "Excel file imported successfully.")

        except Exception as e:
            messages.error(request, f"Error importing file: {e}")

    return redirect('vitamin:stokobat-list')

