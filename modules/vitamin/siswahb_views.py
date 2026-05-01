import os
from io import BytesIO

import pandas as pd
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.models import Group
from django.http import FileResponse, Http404
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.db.models import Q
from django.views.generic import DetailView, ListView
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from openpyxl import load_workbook

from modules.uman.decorators import group_must_have_permission
from modules.vitamin.models import Puskesmas, Sekolah, Siswa, SiswaHB
from modules.vitamin.utils.hb_rag_export import export_siswahb_queryset_to_chroma

from .forms import SiswaHbForm


@method_decorator([group_must_have_permission], name="dispatch")
class SiswaHbListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = SiswaHB
    permission_required = "vitamin.view_siswahb"

    def get_template_names(self):
        if self.request.htmx:
            return ["vitamin/sekolah/siswahb/siswahb_list.html"]
        return ["vitamin/sekolah/siswahb/siswahb.html"]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Data HB Siswa"
        context["q"] = self.request.GET.get("q", "")

        try:
            app_group = Group.objects.get(name="administrator")
            is_admin = self.request.user.is_superuser or app_group in self.request.user.groups.all()
        except Group.DoesNotExist:
            is_admin = self.request.user.is_superuser

        context["is_admin_user"] = is_admin
        return context

    def get_queryset(self):
        search_query = self.request.GET.get("q", "").strip()

        try:
            app_group = Group.objects.get(name="administrator")
            is_admin = self.request.user.is_superuser or app_group in self.request.user.groups.all()
        except Group.DoesNotExist:
            is_admin = self.request.user.is_superuser

        qs = self.model.objects.select_related("siswa", "siswa__sekolah").all()

        if not is_admin:
            if self.request.user.groups.filter(name="Sekolah").exists():
                try:
                    sekolah_id = Sekolah.objects.get(kode=self.request.user.username).id
                    qs = qs.filter(siswa__sekolah_id=sekolah_id)
                except Sekolah.DoesNotExist:
                    qs = qs.none()
            elif self.request.user.groups.filter(name="Puskesmas").exists():
                try:
                    puskesmas = Puskesmas.objects.get(kode=self.request.user.username)
                    sekolah_ids = Sekolah.objects.filter(puskesmas=puskesmas).values_list("id", flat=True)
                    qs = qs.filter(siswa__sekolah_id__in=sekolah_ids)
                except Puskesmas.DoesNotExist:
                    qs = qs.none()
            else:
                qs = qs.none()

        if search_query:
            qs = qs.filter(
                Q(siswa__nis__icontains=search_query)
                | Q(siswa__nama__icontains=search_query)
                | Q(siswa__sekolah__nama__icontains=search_query)
                | Q(tahun__icontains=search_query)
                | Q(keterangan__icontains=search_query)
            )

        return qs.order_by("-id")


class SiswaHbCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = SiswaHB
    template_name = "vitamin/sekolah/siswahb/siswahb_form.html"
    form_class = SiswaHbForm
    success_url = reverse_lazy("vitamin:siswahb-list")
    permission_required = "vitamin.add_siswahb"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Tambah Data HB Siswa"
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.created_by = self.request.user.id
        obj.created_at = timezone.now()
        obj.save()
        return super().form_valid(form)


class SiswaHbUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = SiswaHB
    template_name = "vitamin/sekolah/siswahb/siswahb_edit.html"
    form_class = SiswaHbForm
    success_url = reverse_lazy("vitamin:siswahb-list")
    permission_required = "vitamin.change_siswahb"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Ubah Data HB Siswa"
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.updated_by = self.request.user.id
        obj.updated_at = timezone.now()
        obj.save()
        return super().form_valid(form)


class SiswaHbDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = SiswaHB
    template_name = "vitamin/sekolah/siswahb/siswahb_delete.html"
    success_url = reverse_lazy("vitamin:siswahb-list")
    permission_required = "vitamin.delete_siswahb"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Hapus Data HB Siswa"
        return context


class SiswaHbDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = SiswaHB
    template_name = "vitamin/sekolah/siswahb/siswahb_detail.html"
    success_url = reverse_lazy("vitamin:siswahb-list")
    permission_required = "vitamin.view_siswahb"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Detail Data HB Siswa"
        return context


def download_template(request):
    template_path = os.path.join(settings.BASE_DIR, "static", "files", "template_siswa_hb.xlsx")
    if not os.path.exists(template_path):
        raise Http404("Template file not found.")

    wb = load_workbook(template_path)
    siswa_sheet = "siswa"
    if siswa_sheet not in wb.sheetnames:
        raise Http404(f"Sheet '{siswa_sheet}' not found in template.")
    ws_siswa = wb[siswa_sheet]

    for row in ws_siswa.iter_rows(min_row=2, max_row=ws_siswa.max_row):
        for cell in row:
            cell.value = None

    try:
        app_group = Group.objects.get(name="administrator")
        is_admin = request.user.is_superuser or app_group in request.user.groups.all()
    except Group.DoesNotExist:
        is_admin = request.user.is_superuser

    if is_admin:
        siswa_data = Siswa.objects.select_related("sekolah").values_list("id", "nis", "nama", "sekolah__nama")
    elif request.user.groups.filter(name="Puskesmas").exists():
        try:
            puskesmas = Puskesmas.objects.get(kode=request.user.username)
            sekolah_ids = Sekolah.objects.filter(puskesmas=puskesmas).values_list("id", flat=True)
            siswa_data = Siswa.objects.filter(sekolah_id__in=sekolah_ids).select_related("sekolah").values_list(
                "id", "nis", "nama", "sekolah__nama"
            )
        except Puskesmas.DoesNotExist:
            messages.error(request, "Puskesmas not found for the current user.")
            return redirect("vitamin:siswahb-list")
    else:
        try:
            sekolah = Sekolah.objects.get(kode=request.user.username)
            siswa_data = Siswa.objects.filter(sekolah=sekolah).select_related("sekolah").values_list(
                "id", "nis", "nama", "sekolah__nama"
            )
        except Sekolah.DoesNotExist:
            messages.error(request, "Sekolah not found for the current user.")
            return redirect("vitamin:siswahb-list")

    for idx, (sid, nis, nama, sekolah_nama) in enumerate(siswa_data, start=2):
        ws_siswa.cell(row=idx, column=1).value = sid
        ws_siswa.cell(row=idx, column=2).value = nis
        ws_siswa.cell(row=idx, column=3).value = nama
        ws_siswa.cell(row=idx, column=4).value = sekolah_nama

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return FileResponse(output, as_attachment=True, filename="template_siswa_hb.xlsx")


def import_excel(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]

        try:
            df = pd.read_excel(excel_file, sheet_name="siswa_hb")
            required_columns = {"Id Siswa", "Tahun", "Hb", "Keterangan"}
            if not required_columns.issubset(df.columns):
                missing = required_columns - set(df.columns)
                messages.error(request, f"Missing columns in Excel file: {', '.join(missing)}")
                return redirect("vitamin:siswahb-list")

            try:
                app_group = Group.objects.get(name="administrator")
                is_admin = request.user.is_superuser or app_group in request.user.groups.all()
            except Group.DoesNotExist:
                is_admin = request.user.is_superuser

            sekolah_ids = None
            if not is_admin:
                if request.user.groups.filter(name="Puskesmas").exists():
                    puskesmas = Puskesmas.objects.get(kode=request.user.username)
                    sekolah_ids = set(Sekolah.objects.filter(puskesmas=puskesmas).values_list("id", flat=True))
                else:
                    sekolah = Sekolah.objects.get(kode=request.user.username)
                    sekolah_ids = {sekolah.id}

            for _, row in df.iterrows():
                if pd.isna(row.get("Id Siswa")):
                    continue

                try:
                    siswa = Siswa.objects.select_related("sekolah").get(pk=int(row["Id Siswa"]))
                except Siswa.DoesNotExist:
                    messages.warning(request, f"Siswa with ID {row['Id Siswa']} not found.")
                    continue

                if sekolah_ids is not None and siswa.sekolah_id not in sekolah_ids:
                    messages.warning(request, f"Siswa ID {siswa.id} bukan milik sekolah Anda.")
                    continue

                tahun = int(row["Tahun"]) if not pd.isna(row["Tahun"]) else None
                hb_val = float(row["Hb"]) if not pd.isna(row["Hb"]) else None
                keterangan = str(row["Keterangan"]) if not pd.isna(row["Keterangan"]) else ""

                if tahun is None or hb_val is None:
                    messages.warning(request, "Tahun/Hb kosong, baris dilewati.")
                    continue

                SiswaHB.objects.create(
                    siswa=siswa,
                    tahun=tahun,
                    hb=hb_val,
                    keterangan=keterangan,
                    created_by=request.user.id,
                    created_at=timezone.now(),
                )

            messages.success(request, "Excel file imported successfully.")
        except Exception as e:
            messages.error(request, f"Error importing file: {e}")

    return redirect("vitamin:siswahb-list")


@login_required
def export_hb_to_rag(request):
    try:
        app_group = Group.objects.get(name="administrator")
        is_admin = request.user.is_superuser or app_group in request.user.groups.all()
    except Group.DoesNotExist:
        is_admin = request.user.is_superuser

    if not is_admin:
        messages.error(request, "Akses ditolak.")
        return redirect("vitamin:siswahb-list")

    try:
        qs = SiswaHB.objects.all()
        count = export_siswahb_queryset_to_chroma(qs)
        messages.success(request, f"Sync HB ke RAG berhasil. Total: {count} data.")
    except Exception as e:
        messages.error(request, f"Gagal sync ke RAG: {e}")

    return redirect("vitamin:siswahb-list")
